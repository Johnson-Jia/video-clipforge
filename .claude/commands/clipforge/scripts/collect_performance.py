#!/usr/bin/env python3
"""
视频播放数据自动采集 — 从各平台导出文件中解析播放数据，
匹配到对应的项目目录，产出标准化的 performance 数据。

用法:
  python scripts/collect_performance.py --scan --date 2026-05-29
  python scripts/collect_performance.py --scan --date 2026-05-29 --backfill
  python scripts/collect_performance.py --scan --dry-run
  python scripts/collect_performance.py --scan --json

退出码: 0=成功 1=失败(无文件/无匹配)
"""
from __future__ import annotations

import argparse
import csv
import io
import json
import re
import sys
from datetime import datetime, timezone
from difflib import SequenceMatcher
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

SCRIPT_DIR = Path(__file__).resolve().parent
CLIPFORGE_DIR = SCRIPT_DIR.parent
sys.path.insert(0, str(CLIPFORGE_DIR))
from engine.lib.data_paths import WORKSPACE_ROOT as PROJECT_ROOT  # 四级回退(env>git>config>cwd)
WORKSPACE_ROOT = PROJECT_ROOT / "workspace"


# ── 辅助解析函数 ──────────────────────────────────────────────────────────────

def parse_percent(s) -> float | None:
    """'12.24%' -> 0.1224, '<0.1%' -> 0.001, '-' -> None"""
    if s is None:
        return None
    s = str(s).strip()
    if s in ("-", "", "--"):
        return None
    if "%" in s:
        num = s.replace("%", "").strip()
        if num.startswith("<"):
            try:
                return float(num[1:]) / 100
            except ValueError:
                return None
        if num.startswith(">"):
            try:
                return float(num[1:]) / 100
            except ValueError:
                return None
        try:
            return float(num) / 100
        except ValueError:
            return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def parse_seconds(s) -> float | None:
    """'21.46秒' -> 21.46"""
    if s is None:
        return None
    s = str(s).strip()
    s = s.replace("秒", "")
    if s in ("-", "", "--"):
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def parse_optional_float(s) -> float | None:
    if s is None:
        return None
    s = str(s).strip()
    if s in ("-", "", "--"):
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def parse_bilibili_date(s) -> str | None:
    """'2026年05月29日 12:40:34' -> '2026-05-29'"""
    if not s:
        return None
    m = re.match(r"(\d{4})年(\d{2})月(\d{2})日", str(s))
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    return None


def parse_wechat_date(s) -> str | None:
    """'2026/5/29' -> '2026-05-29'"""
    if not s:
        return None
    try:
        return datetime.strptime(str(s).strip(), "%Y/%m/%d").strftime("%Y-%m-%d")
    except ValueError:
        return None


def parse_xhs_date(s) -> str | None:
    """'2026年05月19日15时39分28秒' -> '2026-05-19'"""
    if not s:
        return None
    m = re.match(r"(\d{4})年(\d{2})月(\d{2})日", str(s))
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    return None


def safe_int(s) -> int:
    if s is None:
        return 0
    try:
        return int(float(str(s).strip().replace(",", "")))
    except (ValueError, TypeError):
        return 0


# ── 平台解析器 ─────────────────────────────────────────────────────────────────

def parse_douyin_xlsx(filepath: Path) -> list[dict]:
    """解析抖音作品列表.xlsx"""
    import openpyxl
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    records = []
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        title_full = str(row[0])
        title = title_full.split("\n")[0].strip()
        plays = safe_int(row[4])
        completion_rate = parse_optional_float(row[5])
        completion_5s = parse_optional_float(row[6])
        likes = safe_int(row[10])
        shares = safe_int(row[11])
        comments = safe_int(row[12])
        saves = safe_int(row[13])
        followers = safe_int(row[15])

        record = {
            "platform": "douyin",
            "title": title,
            "title_full": title_full[:200],
            "published_at": str(row[1])[:19] if row[1] else None,
            "plays": plays,
            "completion_rate": completion_rate,
            "completion_5s_rate": completion_5s,
            "cover_ctr": parse_optional_float(row[7]),
            "bounce_2s_rate": parse_optional_float(row[8]),
            "avg_watch_duration": parse_optional_float(row[9]),
            "likes": likes,
            "shares": shares,
            "comments": comments,
            "saves": saves,
            "profile_visits": safe_int(row[14]),
            "followers_gained": followers,
        }
        if plays > 0:
            record["share_rate"] = round(shares / plays, 6)
            record["like_rate"] = round(likes / plays, 6)
            record["save_rate"] = round(saves / plays, 6)
            record["comment_rate"] = round(comments / plays, 6)
        records.append(record)
    wb.close()
    return records


def parse_bilibili_csv(filepath: Path) -> list[dict]:
    """解析哔哩哔哩近期稿件对比.csv"""
    records = []
    with open(filepath, encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        header = next(reader, None)
        if not header:
            return []
        for row in reader:
            if not row or not row[0]:
                continue
            plays = safe_int(row[2])
            likes = safe_int(row[17])
            comments = safe_int(row[19]) if len(row) > 19 else 0
            danmaku = safe_int(row[21]) if len(row) > 21 else 0
            saves = safe_int(row[23]) if len(row) > 23 else 0
            coins = safe_int(row[25]) if len(row) > 25 else 0
            shares = safe_int(row[27]) if len(row) > 27 else 0
            followers = safe_int(row[14]) if len(row) > 14 else 0

            record = {
                "platform": "bilibili",
                "title": str(row[0]).strip(),
                "published_at": parse_bilibili_date(row[1]),
                "plays": plays,
                "visitor_play_ratio": parse_percent(row[3]),
                "bounce_3s_rate": parse_percent(row[8]) if len(row) > 8 else None,
                "interaction_rate": parse_percent(row[11]) if len(row) > 11 else None,
                "followers_gained": followers,
                "likes": likes,
                "like_rate": parse_percent(row[18]) if len(row) > 18 else None,
                "comments": comments,
                "comment_rate": parse_percent(row[20]) if len(row) > 20 else None,
                "danmaku": danmaku,
                "saves": saves,
                "save_rate": parse_percent(row[24]) if len(row) > 24 else None,
                "coins": coins,
                "shares": shares,
                "share_rate": parse_percent(row[28]) if len(row) > 28 else None,
                "avg_play_progress": parse_percent(row[29]) if len(row) > 29 else None,
            }
            records.append(record)
    return records


def parse_wechat_csv(filepath: Path) -> list[dict]:
    """解析微信视频号动态数据明细.csv"""
    records = []
    with open(filepath, encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        header = next(reader, None)
        if not header:
            return []
        for row in reader:
            if not row or not row[0]:
                continue
            plays = safe_int(row[5])
            likes = safe_int(row[7])
            comments = safe_int(row[8])
            shares = safe_int(row[9])
            follows = safe_int(row[10])
            reshare = safe_int(row[11]) if len(row) > 11 else 0

            desc = str(row[0])
            title = desc[:60].split("\n")[0].strip()

            record = {
                "platform": "wechat_video",
                "title": title,
                "title_full": desc[:200],
                "video_id": row[1] if len(row) > 1 else None,
                "published_at": parse_wechat_date(row[2]) if len(row) > 2 else None,
                "completion_rate": parse_percent(row[3]) if len(row) > 3 else None,
                "avg_watch_duration": parse_seconds(row[4]) if len(row) > 4 else None,
                "plays": plays,
                "recommendations": safe_int(row[6]) if len(row) > 6 else 0,
                "likes": likes,
                "comments": comments,
                "shares": shares,
                "follows": follows,
                "reshare_moments": reshare,
            }
            if plays > 0:
                record["share_rate"] = round(shares / plays, 6)
                record["like_rate"] = round(likes / plays, 6)
            records.append(record)
    return records


def parse_xiaohongshu_xlsx(filepath: Path) -> list[dict]:
    """解析小红书笔记列表明细表.xlsx（第 0 行免责声明跳过，第 1 行表头）"""
    import openpyxl
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        wb.close()
        return []

    records = []
    for row in rows[2:]:
        if not row or not row[1]:
            continue
        title = str(row[0]).strip() if row[0] else ""
        plays = safe_int(row[4]) if len(row) > 4 else 0
        likes = safe_int(row[6]) if len(row) > 6 else 0
        comments = safe_int(row[7]) if len(row) > 7 else 0
        saves = safe_int(row[8]) if len(row) > 8 else 0
        followers = safe_int(row[9]) if len(row) > 9 else 0
        shares = safe_int(row[10]) if len(row) > 10 else 0

        record = {
            "platform": "xiaohongshu",
            "title": title,
            "published_at": parse_xhs_date(row[1]) if len(row) > 1 else None,
            "type": str(row[2]) if len(row) > 2 and row[2] else None,
            "impressions": safe_int(row[3]) if len(row) > 3 else 0,
            "plays": plays,
            "cover_ctr": parse_optional_float(row[5]) if len(row) > 5 else None,
            "likes": likes,
            "comments": comments,
            "saves": saves,
            "followers_gained": followers,
            "shares": shares,
            "avg_watch_duration": parse_optional_float(row[11]) if len(row) > 11 else None,
        }
        if plays > 0:
            record["save_rate"] = round(saves / plays, 6)
            record["like_rate"] = round(likes / plays, 6)
        if likes > 0:
            record["save_to_like_ratio"] = round(saves / likes, 2)
        records.append(record)
    wb.close()
    return records


def _read_xlsx_raw(filepath: Path) -> list[list[str]]:
    """直接通过 zip/XML 读取 xlsx，绕过 openpyxl 样式表解析。

    头条后台导出的 xlsx 样式表损坏（openpyxl 加载报
    `Fill() takes no arguments`），此函数绕过样式层只取单元格值。
    返回二维字符串列表（行内缺失单元格填 ""）。
    """
    import zipfile
    from xml.etree import ElementTree as ET
    NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    rows_out: list[list[str]] = []
    with zipfile.ZipFile(str(filepath)) as z:
        shared: list[str] = []
        try:
            sst = ET.fromstring(z.read("xl/sharedStrings.xml"))
            for si in sst.findall("{%s}si" % NS):
                shared.append("".join((t.text or "") for t in si.iter("{%s}t" % NS)))
        except KeyError:
            pass
        sheet = ET.fromstring(z.read("xl/worksheets/sheet1.xml"))
        for row in sheet.findall(".//{%s}row" % NS):
            cells: dict[int, str] = {}
            max_idx = -1
            for c in row.findall("{%s}c" % NS):
                ref = c.get("r", "")
                col_letters = "".join(ch for ch in ref if ch.isalpha())
                idx = 0
                for ch in col_letters:
                    idx = idx * 26 + (ord(ch.upper()) - ord("A") + 1)
                idx -= 1
                if idx < 0:
                    continue
                v = c.find("{%s}v" % NS)
                val = v.text if v is not None else ""
                ctype = c.get("t", "")
                if ctype == "s" and val.isdigit():
                    n = int(val)
                    val = shared[n] if n < len(shared) else ""
                elif ctype == "inlineStr":
                    is_elem = c.find("{%s}is" % NS)
                    if is_elem is not None:
                        val = "".join((t.text or "") for t in is_elem.iter("{%s}t" % NS))
                cells[idx] = val
                if idx > max_idx:
                    max_idx = idx
            if max_idx < 0:
                rows_out.append([])
            else:
                rows_out.append([cells.get(i, "") for i in range(max_idx + 1)])
    return rows_out


def parse_toutiao_date(s) -> str | None:
    """头条发布时间 → YYYY-MM-DD。兼容 '2026-06-10 12:34' / '2026年06月10日' / '2026/6/10'。"""
    if not s:
        return None
    s = str(s).strip()
    m = re.match(r"(\d{4})\D(\d{1,2})\D(\d{1,2})", s)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    return None


def _parse_toutiao_percent(s) -> float | None:
    """头条完成率解析。'16.55' 表示 16.55%（裸百分数），'-1.00' 为无数据占位。

    头条同一表格内点击率是 0~1 比例小数（'0.46'=46%，走 parse_percent），
    但完成率是裸百分数（'16.55'=16.96%）——两种格式并存，需分别处理。
    验证：行6 展现11668×播放4755→40.7%，CTR=0.41（比例小数✓）；
    完成率列9=23.37（裸百分数，需/100→0.2337）。
    """
    if s is None:
        return None
    s = str(s).strip()
    if s in ("", "-"):
        return None
    try:
        v = float(s)
    except (ValueError, TypeError):
        return None
    if v < 0:  # -1.00 是头条"无数据"占位
        return None
    return round(v / 100, 6)


def _parse_toutiao_seconds(s) -> float | None:
    """头条平均播放时长解析。'-1' 为无数据占位，否则秒数。"""
    if s is None:
        return None
    s = str(s).strip()
    if s in ("", "-"):
        return None
    try:
        v = float(s)
    except (ValueError, TypeError):
        return None
    if v < 0:
        return None
    return v


def parse_toutiao_xlsx(filepath: Path) -> list[dict]:
    """解析头条小视频导出 xlsx（样式表损坏，走 zip/XML 原始读取）。

    16 列：标题(0) / 发布时间(1) / ID(2) / 链接(3) / 展现量(4) / 粉丝展现量(5) /
    播放量(6) / 粉丝播放量(7) / 点击率(8) / 平均播放完成率(9) / 平均播放时长(10) /
    点赞量(11) / 评论量(12) / 转发量(13) / 分享量(14) / 收藏量(15)

    展现量(impressions) 是头条特有指标——平台推送曝光量，区别于播放量。
    """
    rows = _read_xlsx_raw(filepath)
    if len(rows) < 2:
        return []

    records = []
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        title = str(row[0]).strip()
        plays = safe_int(row[6]) if len(row) > 6 else 0
        impressions = safe_int(row[4]) if len(row) > 4 else 0
        likes = safe_int(row[11]) if len(row) > 11 else 0
        comments = safe_int(row[12]) if len(row) > 12 else 0
        forwards = safe_int(row[13]) if len(row) > 13 else 0
        shares_col = safe_int(row[14]) if len(row) > 14 else 0
        saves = safe_int(row[15]) if len(row) > 15 else 0

        record = {
            "platform": "toutiao",
            "title": title,
            "published_at": parse_toutiao_date(row[1]) if len(row) > 1 else None,
            "impressions": impressions,
            "plays": plays,
            "click_rate": parse_percent(row[8]) if len(row) > 8 else None,
            "completion_rate": _parse_toutiao_percent(row[9]) if len(row) > 9 else None,
            "avg_watch_duration": _parse_toutiao_seconds(row[10]) if len(row) > 10 else None,
            "likes": likes,
            "comments": comments,
            "forwards": forwards,
            "shares": shares_col + forwards,  # 转发+分享合计，对齐既有 share_rate 语义
            "saves": saves,
        }
        if plays > 0:
            record["share_rate"] = round(record["shares"] / plays, 6)
            record["like_rate"] = round(likes / plays, 6)
            record["save_rate"] = round(saves / plays, 6)
        records.append(record)
    return records


# ── 数据文件扫描 ───────────────────────────────────────────────────────────────

def scan_data_dir(data_root: Path, date_filter: str | None = None) -> list[dict]:
    """扫描数据目录，返回所有平台记录。"""
    all_records: list[dict] = []
    data_root = Path(data_root)

    if not data_root.exists():
        return []

    date_dirs = sorted(data_root.iterdir()) if data_root.is_dir() else []
    for dd in date_dirs:
        if not dd.is_dir():
            continue
        date_str = dd.name
        if date_filter and date_str != date_filter:
            continue

        # 抖音
        douyin_file = dd / "抖音作品列表.xlsx"
        if douyin_file.exists():
            recs = parse_douyin_xlsx(douyin_file)
            for r in recs:
                r["_data_date"] = date_str
            all_records.extend(recs)

        # B站（导出按每页10条分多个文件：b站近期稿件对比1.csv ~ N.csv，
        # 命名/大小写不一，旧版用「哔哩哔哩」前缀，glob 兼容）
        for bl_file in sorted(dd.glob("*.csv")):
            fn = bl_file.name
            if ("稿件对比" not in fn) or ("哔哩哔哩" not in fn and "b站" not in fn.lower()):
                continue
            recs = parse_bilibili_csv(bl_file)
            for r in recs:
                r["_data_date"] = date_str
            all_records.extend(recs)

        # 视频号（两个可能的文件名）
        for fname in ("微信视频号动态数据明细.csv", "视频号动态数据明细.csv"):
            wechat_file = dd / fname
            if wechat_file.exists():
                recs = parse_wechat_csv(wechat_file)
                for r in recs:
                    r["_data_date"] = date_str
                all_records.extend(recs)
                break

        # 小红书
        xhs_file = dd / "小红书笔记列表明细表.xlsx"
        if xhs_file.exists():
            recs = parse_xiaohongshu_xlsx(xhs_file)
            for r in recs:
                r["_data_date"] = date_str
            all_records.extend(recs)

        # 头条（前缀匹配，兼容 "头条小视频*.xlsx" / "头条-小视频*.xlsx"）
        for tt_file in sorted(dd.glob("头条*.xlsx")):
            recs = parse_toutiao_xlsx(tt_file)
            for r in recs:
                r["_data_date"] = date_str
            all_records.extend(recs)
            break  # 每个日期目录只取一个头条导出文件

    return all_records


# ── 项目索引与匹配 ─────────────────────────────────────────────────────────────

def _extract_titles_from_douyin_md(content: str) -> list[str]:
    """从 douyin.md 中提取所有文案标题（支持多种格式）。"""
    titles = []
    # 格式 1: "## 抖音" section 下的第一个非空行
    in_section = False
    for line in content.split("\n"):
        stripped = line.strip()
        if stripped == "## 抖音":
            in_section = True
            continue
        if in_section:
            if stripped.startswith("## ") or stripped == "---":
                in_section = False
                continue
            if stripped and not stripped.startswith("#"):
                titles.append(stripped)
                break
    # 格式 2: "## 文案 N" sections
    if not titles:
        for line in content.split("\n"):
            stripped = line.strip()
            if stripped.startswith("## 文案") or stripped.startswith("## 标题"):
                in_section = True
                continue
            if in_section:
                if stripped.startswith("## ") or stripped == "---":
                    in_section = False
                    continue
                if stripped and not stripped.startswith("#"):
                    titles.append(stripped)
    return titles


def build_project_index(workspace_root: Path) -> list[dict]:
    """构建项目索引：基于 final.mp4/output.mp4 存在的项目目录。"""
    workspace_root = Path(workspace_root)
    projects = []
    seen_dirs = set()

    # 第一轮：有 douyin.md 的项目
    for dm in sorted(workspace_root.rglob("douyin.md")):
        project_dir = dm.parent
        rel = project_dir.relative_to(workspace_root)
        parts = rel.parts
        if "test" in parts:
            continue  # test 目录隔离，不进 collect 数据
        if len(parts) < 3:
            continue
        date_path = f"{parts[0]}/{parts[1]}/{parts[2]}"
        project_name = parts[3] if len(parts) > 3 else parts[2]

        titles = []
        try:
            content = dm.read_text(encoding="utf-8")
            titles = _extract_titles_from_douyin_md(content)
        except (UnicodeDecodeError, OSError):
            pass

        # 备用：narration.txt 首行
        narration_first = ""
        nf = project_dir / "narration.txt"
        if nf.exists():
            try:
                narration_first = nf.read_text(encoding="utf-8").split("\n")[0].strip()
            except (UnicodeDecodeError, OSError):
                pass

        projects.append({
            "dir": project_dir,
            "rel_path": str(rel),
            "date_path": date_path,
            "project_name": project_name,
            "douyin_title": titles[0] if titles else "",
            "douyin_titles": titles,
            "narration_first": narration_first,
        })
        seen_dirs.add(str(project_dir))

    # 第二轮：有 final.mp4 但没有 douyin.md 的项目
    for fm in sorted(workspace_root.rglob("final.mp4")):
        project_dir = fm.parent
        if str(project_dir) in seen_dirs:
            continue
        rel = project_dir.relative_to(workspace_root)
        parts = rel.parts
        if "test" in parts:
            continue  # test 目录隔离，不进 collect 数据
        if len(parts) < 3:
            continue
        date_path = f"{parts[0]}/{parts[1]}/{parts[2]}"
        project_name = parts[3] if len(parts) > 3 else parts[2]

        narration_first = ""
        nf = project_dir / "narration.txt"
        if nf.exists():
            try:
                narration_first = nf.read_text(encoding="utf-8").split("\n")[0].strip()
            except (UnicodeDecodeError, OSError):
                pass

        projects.append({
            "dir": project_dir,
            "rel_path": str(rel),
            "date_path": date_path,
            "project_name": project_name,
            "douyin_title": "",
            "douyin_titles": [],
            "narration_first": narration_first,
        })

    return projects


def _build_date_project_index(projects: list[dict]) -> dict[str, list[dict]]:
    """构建日期 → 项目索引，供空标题匹配使用。"""
    by_date: dict[str, list[dict]] = {}
    for p in projects:
        dp = p.get("date_path", "")
        if dp:
            by_date.setdefault(dp, []).append(p)
    return by_date


def _build_date_record_ranks(records: list[dict], platform: str) -> dict[str, list[tuple[int, dict]]]:
    """按日期对同平台记录按播放量排名，返回 {date: [(rank, record), ...]}。"""
    by_date: dict[str, list[dict]] = {}
    for r in records:
        if r.get("platform") != platform or not r.get("published_at"):
            continue
        dp = r["published_at"][:10]
        by_date.setdefault(dp, []).append(r)
    ranks: dict[str, list[tuple[int, dict]]] = {}
    for dp, recs in by_date.items():
        recs.sort(key=lambda x: x.get("plays", 0), reverse=True)
        ranks[dp] = [(i + 1, r) for i, r in enumerate(recs)]
    return ranks


def _match_empty_title(
    record: dict,
    date_projects_index: dict[str, list[dict]],
    douyin_date_ranks: dict[str, list[tuple[int, dict]]],
) -> tuple[dict | None, float, float]:
    """空标题记录的日期匹配。

    策略：同日唯一项目直接匹配；多项目无法消歧则跳过（宁缺毋误）。
    """
    rec_date = record.get("published_at", "")[:10]
    dp = rec_date.replace("-", "/")
    candidates = date_projects_index.get(dp, [])
    if not candidates:
        return None, 0.0, 0.0

    # 同日唯一项目 → 高置信匹配
    if len(candidates) == 1:
        return candidates[0], 0.55, 0.0

    # 多项目：无标题无法消歧，跳过
    return None, 0.0, 0.0


def match_records_to_projects(
    records: list[dict],
    projects: list[dict],
) -> list[dict]:
    """将记录匹配到项目。返回 [{record, project, score}]。

    匹配信号:
      1. 标题相似度 — 取所有文案标题/narration首行的最佳匹配 (权重 0.5)
      2. 日期精确匹配 (权重 0.3)
      3. 关键词重叠 (权重 0.2)
      4. 空标题时：日期+排名匹配（独立路径）
    """
    matches = []
    # 预构建索引（空标题匹配用）
    date_project_idx = _build_date_project_index(projects)
    xhs_date_ranks = _build_date_record_ranks(records, "xiaohongshu")

    for record in records:
        best_project = None
        best_score = 0.0
        second_score = 0.0

        rec_title = record.get("title", "")
        rec_platform = record.get("platform", "")
        rec_date = record.get("published_at", "")

        # 空标题快速路径：日期+排名匹配
        if not rec_title.strip():
            proj, score, second = _match_empty_title(
                record, date_project_idx, xhs_date_ranks,
            )
            matched = proj is not None and score >= 0.40 and (score - second >= 0.10 or score >= 0.55)
            matches.append({
                "record": record,
                "project": proj if matched else None,
                "score": round(score, 3),
                "matched": matched,
            })
            continue

        # 从发布日期提取 YYYY/MM/DD
        date_parts = None
        dp = None
        if rec_date:
            dp = rec_date[:10].split("-")
            if len(dp) == 3:
                date_parts = f"{dp[0]}/{dp[1]}/{dp[2]}"

        for proj in projects:
            score = 0.0

            # 信号 1: 标题相似度 — 多源取最佳 (权重 0.5)
            best_ratio = 0.0
            if rec_title:
                candidates = list(proj.get("douyin_titles", []))
                if proj.get("douyin_title") and proj["douyin_title"] not in candidates:
                    candidates.insert(0, proj["douyin_title"])
                if proj.get("narration_first"):
                    candidates.append(proj["narration_first"])
                for cand in candidates:
                    if not cand:
                        continue
                    ratio = SequenceMatcher(
                        None,
                        rec_title[:80].lower(),
                        cand[:80].lower(),
                    ).ratio()
                    if ratio > best_ratio:
                        best_ratio = ratio
            if best_ratio > 0:
                score += 0.5 * best_ratio
            elif rec_title:
                proj_name_words = proj["project_name"].replace("-", " ").lower().split()
                title_lower = rec_title.lower()
                overlap = sum(1 for w in proj_name_words if len(w) > 2 and w in title_lower)
                if proj_name_words:
                    score += 0.5 * min(overlap / max(len(proj_name_words), 1), 1.0)

            # 信号 2: 日期目录匹配 (权重 0.3)
            if date_parts:
                if date_parts == proj["date_path"]:
                    score += 0.3
                try:
                    from datetime import date, timedelta
                    y, m, d = map(int, dp)
                    rec_d = date(y, m, d)
                    pp = proj["date_path"].split("/")
                    proj_d = date(int(pp[0]), int(pp[1]), int(pp[2]))
                    delta = abs((rec_d - proj_d).days)
                    if delta == 1:
                        score += 0.15
                    elif delta == 2:
                        score += 0.05
                except (ValueError, IndexError):
                    pass

            # 信号 3: 关键词重叠 — 从所有文案标题汇总关键词 (权重 0.2)
            if rec_title:
                proj_words = set(proj["project_name"].replace("-", " ").lower().split())
                for t in proj.get("douyin_titles", []):
                    proj_words.update(t[:80].lower().split())
                if proj.get("douyin_title"):
                    proj_words.update(proj["douyin_title"][:80].lower().split())
                if proj.get("narration_first"):
                    proj_words.update(proj["narration_first"][:80].lower().split())
                rec_words = set(rec_title[:80].lower().split())
                stop_words = {"的", "了", "是", "在", "和", "与", "及", "到", "有", "不", "这", "那", "一", "个"}
                proj_words -= stop_words
                rec_words -= stop_words
                proj_words = {w for w in proj_words if len(w) >= 2}
                rec_words = {w for w in rec_words if len(w) >= 2}
                if proj_words and rec_words:
                    overlap = len(proj_words & rec_words) / max(len(proj_words), 1)
                    score += 0.2 * min(overlap, 1.0)

            if score > best_score:
                second_score = best_score
                best_score = score
                best_project = proj
            elif score > second_score:
                second_score = score

        # 匹配判定
        matched = False
        if best_score >= 0.4 and best_project:
            if best_score - second_score >= 0.10:
                matched = True
            elif best_score >= 0.6:
                matched = True

        matches.append({
            "record": record,
            "project": best_project if matched else None,
            "score": round(best_score, 3),
            "matched": matched,
        })

    return matches


# ── 回填逻辑 ──────────────────────────────────────────────────────────────────

def backfill_matches(
    matches: list[dict],
    workspace_root: Path,
) -> list[dict]:
    """执行回填：写入 performance.json 并调用 trace.backfill。"""
    results = []

    # 按项目分组
    by_project: dict[str, list[dict]] = {}
    for m in matches:
        if not m["matched"] or not m["project"]:
            continue
        key = str(m["project"]["dir"])
        by_project.setdefault(key, []).append(m)

    # 导入 trace 模块
    try:
        sys.path.insert(0, str(CLIPFORGE_DIR))
        from engine.trace import backfill_performance
        has_trace = True
    except ImportError:
        has_trace = False

    for proj_dir_str, group in by_project.items():
        proj_dir = Path(proj_dir_str)
        platforms = {}
        new_snapshots = []

        for m in group:
            rec = m["record"]
            platform = rec["platform"]
            # 清理内部字段
            perf = {k: v for k, v in rec.items() if not k.startswith("_") and v is not None}
            platforms[platform] = perf
            # 构建快照
            snapshot_fields = {"plays", "likes", "comments", "shares", "saves",
                               "completion_5s_rate", "completion_rate", "cover_ctr",
                               "share_rate", "like_rate", "save_rate",
                               "followers_gained", "avg_watch_duration",
                               "impressions", "click_rate", "forwards"}
            snap = {"platform": platform, "source_date": rec.get("_data_date", "")}
            for f in snapshot_fields:
                if f in perf:
                    snap[f] = perf[f]
            new_snapshots.append(snap)

        if not platforms:
            continue

        # 合并历史：读取已有 performance.json，保留 snapshots
        perf_path = proj_dir / "performance.json"
        existing_snapshots = []
        if perf_path.exists():
            try:
                old = json.loads(perf_path.read_text(encoding="utf-8"))
                existing_snapshots = old.get("snapshots", [])
            except Exception:
                pass

        # 去重：同 source_date + platform 的旧快照被新数据替换
        new_keys = {(s["platform"], s["source_date"]) for s in new_snapshots}
        merged = [s for s in existing_snapshots
                  if (s.get("platform"), s.get("source_date")) not in new_keys]
        merged.extend(new_snapshots)
        # 按时间排序
        merged.sort(key=lambda s: s.get("source_date", ""))

        # 写入 performance.json（保留历史快照）
        perf_data = {
            "project": str(proj_dir.relative_to(workspace_root)),
            "collected_at": datetime.now(timezone.utc).isoformat(),
            "data_sources": list({m["record"]["_data_date"] for m in group}),
            "platforms": platforms,
            "snapshots": merged,
        }
        perf_path.write_text(json.dumps(perf_data, ensure_ascii=False, indent=2), encoding="utf-8")

        # 回填到 trace
        trace_updated = 0
        if has_trace:
            for platform, perf in platforms.items():
                n = backfill_performance(str(proj_dir), perf)
                trace_updated += n

        results.append({
            "project_dir": str(proj_dir.relative_to(workspace_root)),
            "platforms": list(platforms.keys()),
            "trace_updated": trace_updated,
        })

    return results


def run_feedback_loop(
    matches: list[dict],
    workspace_root: Path,
) -> tuple[list[dict], dict | None]:
    """Post-backfill feedback chain: attribution → calibration → pattern extraction.

    反馈循环的激活点 — 当播放数据到达并回填后自动触发。
    """
    try:
        sys.path.insert(0, str(CLIPFORGE_DIR))
        from engine.attribution import performance_attribution, calibrate_machine_scoring, _classify_hook_type
        from engine.success_analyzer import auto_extract_from_performance, save_pattern
    except ImportError as e:
        return [], {"error": f"Import failed: {e}"}

    feedback_results = []
    all_perf_data = []

    for m in matches:
        if not m["matched"] or not m["project"]:
            continue

        proj_dir = Path(m["project"]["dir"])
        rec = m["record"]
        perf = {k: v for k, v in rec.items() if not k.startswith("_") and v is not None}
        platform = rec.get("platform", "")
        narration_file = proj_dir / "narration_segments.json"

        # Hook type classification for pattern extraction
        if narration_file.exists():
            try:
                seg_data = json.loads(narration_file.read_text(encoding="utf-8"))
                segments = seg_data.get("segments", [])
                if segments:
                    hook_text = segments[0].get("narration_segment", "")
                    perf["hook_type"] = _classify_hook_type(hook_text)
            except Exception:
                pass

        # 1. Performance attribution → Delta
        attr_summary = None
        try:
            attr = performance_attribution(perf, narration_file)
            attr_summary = {
                "root_cause": attr.get("root_cause"),
                "confidence": attr.get("confidence"),
                "causes": [c.get("cause") for c in attr.get("causes", [])],
                "delta_path": attr.get("delta_path"),
            }
        except Exception as e:
            attr_summary = {"error": str(e)}

        # 2. Machine scoring calibration (if score_report.json exists)
        cal_summary = None
        score_report_path = proj_dir / "score_report.json"
        if score_report_path.exists():
            try:
                score_report = json.loads(score_report_path.read_text(encoding="utf-8"))
                cal = calibrate_machine_scoring(
                    score_report, perf, narration_file=narration_file
                )
                cal_summary = {
                    "verdict": cal.get("verdict"),
                    "prediction": cal.get("machine_prediction"),
                    "outcome": cal.get("actual_outcome"),
                    "diagnosis": cal.get("diagnosis"),
                    "delta_path": cal.get("delta_path"),
                    "freshness_signal": cal.get("freshness_signal"),
                }
            except Exception as e:
                cal_summary = {"error": str(e)}

        all_perf_data.append(perf)
        feedback_results.append({
            "project_dir": str(proj_dir.relative_to(workspace_root)),
            "platform": platform,
            "attribution": attr_summary,
            "calibration": cal_summary,
        })

    # 3. Pattern extraction (needs >= 3 samples)
    pattern_summary = None
    if len(all_perf_data) >= 3:
        try:
            patterns = auto_extract_from_performance(all_perf_data)
            if patterns:
                saved_paths = []
                for p in patterns:
                    path = save_pattern(p)
                    saved_paths.append(str(path))
                pattern_summary = {
                    "count": len(patterns),
                    "pattern_ids": [p["id"] for p in patterns],
                    "saved": saved_paths,
                }
        except Exception:
            pass

    return feedback_results, pattern_summary


# ── 报告输出 ──────────────────────────────────────────────────────────────────

def print_table_report(matches: list[dict], backfill_results: list[dict] | None):
    """打印表格形式的报告。"""
    # 按平台分组
    by_platform: dict[str, list[dict]] = {}
    for m in matches:
        p = m["record"]["platform"]
        by_platform.setdefault(p, []).append(m)

    platform_names = {
        "douyin": "抖音", "bilibili": "B站",
        "wechat_video": "视频号", "xiaohongshu": "小红书",
        "toutiao": "头条",
    }

    for platform, group in by_platform.items():
        name = platform_names.get(platform, platform)
        matched_count = sum(1 for m in group if m["matched"])
        print(f"\n[{name}] {len(group)} 条记录, {matched_count} 条匹配")

        for m in group:
            rec = m["record"]
            title = rec.get("title", "")[:40]
            plays = rec.get("plays", 0)

            if m["matched"]:
                proj = m["project"]
                print(f"  MATCH {proj['rel_path']} ← \"{title}\" ({m['score']:.2f})")
                # 打印关键指标
                metrics = []
                if platform == "douyin":
                    c5s = rec.get("completion_5s_rate")
                    if c5s is not None:
                        metrics.append(f"5s完播={c5s:.1%}")
                    comp = rec.get("completion_rate")
                    if comp is not None:
                        metrics.append(f"完播={comp:.1%}")
                elif platform == "bilibili":
                    bounce = rec.get("bounce_3s_rate")
                    if bounce is not None:
                        metrics.append(f"3s跳出={bounce:.1%}")
                elif platform == "wechat_video":
                    comp = rec.get("completion_rate")
                    if comp is not None:
                        metrics.append(f"完播={comp:.1%}")
                    shares = rec.get("shares", 0)
                    if plays > 0:
                        metrics.append(f"分享={shares}({shares/plays:.1%})")
                elif platform == "xiaohongshu":
                    saves = rec.get("saves", 0)
                    likes = rec.get("likes", 0)
                    if likes > 0:
                        metrics.append(f"收藏/赞={saves/likes:.1f}")
                elif platform == "toutiao":
                    metrics.append(f"展现={rec.get('impressions', 0)}")
                    ctr = rec.get("click_rate")
                    if ctr is not None:
                        metrics.append(f"点击率={ctr:.1%}")
                    comp = rec.get("completion_rate")
                    if comp is not None:
                        metrics.append(f"完播={comp:.1%}")

                metrics.insert(0, f"播放={plays}")
                print(f"    {' '.join(metrics)}")
            else:
                print(f"  SKIP  \"{title}\" (best={m['score']:.2f})")

    # 汇总
    total = len(matches)
    matched = sum(1 for m in matches if m["matched"])
    print(f"\n=== 汇总 ===")
    print(f"匹配: {matched}/{total} 条记录 ({len(by_platform)} 平台)")

    if backfill_results:
        print(f"回填: {len(backfill_results)} 个项目")
        for r in backfill_results:
            print(f"  {r['project_dir']}: {', '.join(r['platforms'])} (trace更新{r['trace_updated']}条)")


# ── 主入口 ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="视频播放数据自动采集")
    parser.add_argument("--scan", action="store_true", help="扫描数据目录")
    parser.add_argument("--workspace-root", default=str(WORKSPACE_ROOT), help="workspace 根目录")
    parser.add_argument("--data-root", default=str(WORKSPACE_ROOT / "sources" / "视频数据"), help="视频数据根目录")
    parser.add_argument("--date", default=None, help="只处理指定日期 (YYYY-MM-DD)")
    parser.add_argument("--backfill", action="store_true", help="自动回填到项目目录")
    parser.add_argument("--dry-run", action="store_true", help="显示匹配结果但不写入")
    parser.add_argument("--json", action="store_true", help="JSON 格式输出")
    args = parser.parse_args()

    if not args.scan:
        parser.print_help()
        return 1

    workspace_root = Path(args.workspace_root)
    data_root = Path(args.data_root)

    # 扫描数据
    records = scan_data_dir(data_root, args.date)
    if not records:
        print("未找到数据文件")
        return 1

    # 构建项目索引
    projects = build_project_index(workspace_root)
    if not projects:
        print("未找到视频项目")
        return 1

    # 匹配
    matches = match_records_to_projects(records, projects)

    # 回填
    backfill_results = None
    feedback_results = None
    pattern_results = None
    if args.backfill and not args.dry_run:
        backfill_results = backfill_matches(matches, workspace_root)
        # 反馈循环: backfill 后自动触发 attribution → calibration → pattern extraction
        if backfill_results:
            try:
                feedback_results, pattern_results = run_feedback_loop(matches, workspace_root)
            except Exception:
                pass

    # 输出
    if args.json:
        output = {
            "total_records": len(records),
            "total_matches": sum(1 for m in matches if m["matched"]),
            "matches": [],
            "unmatched": [],
        }
        for m in matches:
            entry = {
                "platform": m["record"]["platform"],
                "title": m["record"].get("title", "")[:60],
                "score": m["score"],
                "matched": m["matched"],
            }
            if m["matched"]:
                entry["project_dir"] = m["project"]["rel_path"]
                entry["performance"] = {k: v for k, v in m["record"].items()
                                        if not k.startswith("_") and v is not None}
            output["matches" if m["matched"] else "unmatched"].append(entry)

        if backfill_results:
            output["backfill"] = backfill_results

        if feedback_results:
            output["feedback"] = feedback_results
        if pattern_results:
            output["patterns"] = pattern_results

        print(json.dumps(output, ensure_ascii=False, indent=2))
    else:
        print_table_report(matches, backfill_results)
        if feedback_results:
            print("\n=== 反馈循环 ===")
            for r in feedback_results:
                attr = r.get("attribution") or {}
                cal = r.get("calibration")
                print(f"  {r['project_dir']} ({r['platform']}):")
                if attr.get("root_cause") and attr["root_cause"] != "no_performance_issue":
                    print(f"    归因: {attr['root_cause']} (置信度 {attr.get('confidence', 0):.2f})")
                    if attr.get("delta_path"):
                        print(f"    Delta: {attr['delta_path']}")
                if cal:
                    print(f"    校准: {cal.get('verdict', 'N/A')} ({cal.get('prediction', '?')} → {cal.get('outcome', '?')})")
                    if cal.get("delta_path"):
                        print(f"    Delta: {cal['delta_path']}")
            if pattern_results:
                print(f"  模式提炼: {pattern_results['count']} 个新模式")
                for pid in pattern_results.get("pattern_ids", []):
                    print(f"    - {pid}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
